from django.shortcuts import render

# Create your views here.
from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from openpyxl import Workbook
from citasmd.forms import CitasMdFormulario
from citasmd.models import CitasMd

# Create your views here.


def agregar_citasmd(request):
    pagina = loader.get_template('agregar_citasmd.html')
    if request.method == 'GET':
        formulario= CitasMdFormulario
    elif request.method == 'POST':
        formulario= CitasMdFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_citasmd(request, idCitasMd):
    pagina= loader.get_template('ver_citasmd.html')
    citasmd= get_object_or_404(CitasMd, pk=idCitasMd)
    mensaje= {'citasmd': citasmd}
    return HttpResponse(pagina.render(mensaje, request))


def editar_citasmd(request, idCitasMd):
    pagina= loader.get_template('editar_citasmd.html')
    citasmd= get_object_or_404(CitasMd, pk=idCitasMd)
    if request.method == 'GET':
        formulario= CitasMdFormulario(instance=citasmd)
    elif request.method == 'POST':
        formulario= CitasMdFormulario(request.POST, instance=citasmd)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    mensaje= {'formulario': formulario}
    return HttpResponse(pagina.render(mensaje,request))

def eliminar_citasmd(request, idCitasMd):
    citasmd = get_object_or_404(CitasMd, pk=idCitasMd)
    if citasmd:
        citasmd.delete()
        return redirect('inicio')

def generar_reporte(request):
        # Obtenemos todas las personas de nuestra base de datos
        citasmds = CitasMd.objects.all()
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'REPORTE DE Citas Medicas'
        # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'ID'
        ws['C3'] = 'FECHA DE LA CITA'
        ws['D3'] = 'ACTIVO'
        ws['E3'] = 'NOMBRE DEL DOCTOR'
        ws['F3'] = 'APELLIDO DEL DOCTOR'
        ws['G3'] = 'ESPECIALIDAD DEL DOCTOR'
        ws['H3'] = 'DIRECCION DEL DOCTOR'
        ws['I3'] = 'ID DEL PACIENTE'
        ws['J3'] = 'NOMBRE DEL PACIENTE'
        ws['K3'] = 'APELLIDO DEL PACIENTE'
        ws['L3'] = 'EDAD DEL PACIENTE'
        ws['M3'] = 'CORREO DEL PACIENTE'
        cont = 4
        # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for citasmd in citasmds:
            ws.cell(row=cont, column=2).value = citasmd.id
            ws.cell(row=cont, column=3).value = citasmd.fecha_cita
            ws.cell(row=cont, column=4).value = citasmd.activo
            ws.cell(row=cont, column=5).value = citasmd.doctor.Doc_nomb
            ws.cell(row=cont, column=6).value = citasmd.doctor.Doc_Apellido
            ws.cell(row=cont, column=7).value = citasmd.doctor.especialidad
            ws.cell(row=cont, column=8).value = citasmd.doctor.Doc_direccion
            ws.cell(row=cont, column=9).value = citasmd.paciente.id_paciente
            ws.cell(row=cont, column=10).value = citasmd.paciente.Pac_nomb
            ws.cell(row=cont, column=11).value = citasmd.paciente.Pac_Apellido
            ws.cell(row=cont, column=12).value = citasmd.paciente.Pac_edad
            ws.cell(row=cont, column=13).value = citasmd.paciente.correo
            cont = cont + 1
        # Establecemos el nombre del archivo
        nombre_archivo = "ReportePersonasExcel.xlsx"
        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response